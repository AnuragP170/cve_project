import openpyxl

def reverse_excel_rows(input_file, output_file):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Read the data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Reverse the data (excluding the header if there is one)
    header = data[0]
    data = data[1:]
    data.reverse()
    data.insert(0, header)

    # Write the reversed data back to a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    for row in data:
        new_sheet.append(row)

    # Save the new workbook
    new_wb.save(output_file)

# Specify the input and output file names
input_file = 'extracted_cve_details.xlsx'  # Change this to your input file
output_file = 'extracted_cve_details_2.xlsx'  # Change this to your desired output file

# Reverse the rows and save to a new file
reverse_excel_rows(input_file, output_file)
print(f"Reversed rows have been saved to {output_file}")
